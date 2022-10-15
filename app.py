import requests 
from lxml import html
import openpyxl
def writedata(data):
	wb=openpyxl.Workbook()
	ws=wb.active
	heading=["name", "id", "Problem Solving", "CPP","Java","Python","Days of Js","Days of Code","Days ofStatistics","Sql","C language","Ruby"]
	ws.append(heading)
	for i in range(len(data)):
		values=[data[i][k] for k in data[i].keys()]
		ws.append(values)
	wb.save("output.xlsx")

details=openpyxl.load_workbook("./input.xlsx")
ws = details.active
data=dict()
finaldata=[]
rows=len([row for row in ws if not all([cell.value is None for cell in row])])
for i in range(2,rows+1):
	print(str(i-1)+"/"+str(rows)+" user completed")
	data["name"]=ws.cell(i,1).value;
	data["id"]=ws.cell(i,2).value;
	data["Problem Solving"]=""
	data["CPP"]=""
	data["Java"]=""
	data["Python"]=""
	data["Days of Js"]=""
	data["Days of Code"]=""
	data["Days ofStatistics"]=""
	data["Sql"]=""
	data["C language"]=""
	data["Ruby"]=""
	url = 'https://www.hackerrank.com/profile/'+ws.cell(row=i,column=2).value
	headers={'User-Agent':"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36"}
	namepath = '/html/body/div[4]/div/div/div/div/div[3]/article/div/div[2]/section[1]/div/div/div/div[1]/div'
	starpath=  '/html/body/div[4]/div/div/div/div/div[3]/article/div/div[2]/section[1]/div/div/div/div[1]/div[1]/div/div/div/svg/g/svg/svg'
	response = requests.get(url,headers=headers)
	byte_data = response.content
	source_code = html.fromstring(byte_data)
	tree = source_code.xpath(namepath)
	for ddata in tree:
		data[ddata.text_content()]=str(len(ddata.xpath('div/div/div/svg/g/svg/svg')))+"★"
	finaldata.append(data.copy())
writedata(finaldata)